from zipfile import ZipFile
import csv
from pypdf import PdfReader
from openpyxl import load_workbook
from conftest import ZIPPED_RESOURCES


# Читаем и проверяем содержимое .csv файла в архиве
def test_read_and_check_csv_file():
    with ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('CSVfile.csv') as csv_file:
            content = csv_file.read().decode('utf-8')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]
            third_row = csvreader[2]

            # Проверяем содержимое второй строки
            assert second_row[1] == '2004.06'

            # Проверяем содержимое третьей строки
            assert third_row[2] == '1220.571'


# Читаем и проверяем содержимое .xlsx файла в архиве
def test_read_and_check_xlsx_file():
    with ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('XLSXfile.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

    # Проверяем количество страниц
    assert len(workbook.sheetnames) == 1

    # Проверяем количество строк
    row_count = sheet.max_row
    assert row_count == 51

    # Проверяем количество столбцов
    column_count = sheet.max_column
    assert column_count == 8

    # Проверяем содержимое пятой строки четвертого столбца
    contents = (sheet.cell(row=5, column=4).value)
    assert contents == "Female"


# Читаем и проверяем содержимое .pdf файла в архиве
def test_read_and_check_pdf_file():
    with ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('PDFfile.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            number_of_pages = len(reader.pages)

    # Проверяем количество страниц в .pdf файле
    assert number_of_pages == 256
